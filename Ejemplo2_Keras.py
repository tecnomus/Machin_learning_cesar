"""
Documentación: Regresión Logística Binaria con Keras y openpyxl

Descripción:
Este script entrena una red neuronal de una sola capa (Regresión Logística). 
Se utiliza TensorFlow estrictamente para convertir las listas nativas de Python 
a 'Tensores', ya que Keras exige este formato de datos para operar.
"""

import openpyxl
import matplotlib.pyplot as plt
import tensorflow as tf # Requisito estricto de Keras para procesar los datos
from keras.models import Sequential
from keras.layers import Dense, Input
from keras.optimizers import SGD

# --- 1. Carga y Preparación de Datos ---
ruta_archivo = 'Regresión.xlsx'

workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
sheet = workbook.active

X_train_lista = [] 
y_train_lista = [] 

for row in range(2, sheet.max_row + 1):
    valor_hora = sheet.cell(row=row, column=1).value
    valor_resultado = sheet.cell(row=row, column=2).value
    
    if valor_hora is not None and valor_resultado is not None:
        try:
            hora_limpia = float(valor_hora)
            resultado_limpio = float(valor_resultado) # Float para compatibilidad con tensores
            
            X_train_lista.append([hora_limpia])
            y_train_lista.append(resultado_limpio)
        except (ValueError, TypeError):
            continue

n_datos = len(X_train_lista)
print(f"Se han cargado {n_datos} registros exitosamente desde el Excel.")

if n_datos == 0:
    print("Error: No se pudieron cargar los datos. Verifica tu archivo Excel.")
    exit()

# CONVERSIÓN OBLIGATORIA A TENSORES: Keras rechaza las listas nativas de Python
X_train = tf.convert_to_tensor(X_train_lista)
y_train = tf.convert_to_tensor(y_train_lista)

# --- 2. Topología de la Red Neuronal ---
model = Sequential()
# Se usa Input() como primera capa para resolver el Warning de Keras
model.add(Input(shape=(1,)))
model.add(Dense(units=1, activation='sigmoid'))

# --- 3. Compilación ---
model.compile(optimizer=SGD(learning_rate=0.05), 
              loss='binary_crossentropy', 
              metrics=['accuracy'])

# --- 4. Entrenamiento Automático ---
print("Iniciando entrenamiento del modelo en Keras...")
historial = model.fit(X_train, y_train, epochs=150, verbose=0) 
print("Entrenamiento finalizado.")

# --- 5. Extracción de Métricas y Parámetros ---
loss_final, precision_final = model.evaluate(X_train, y_train, verbose=0)
print(f"Precisión (Accuracy) sobre los datos de entrenamiento: {precision_final*100:.2f}%")

pesos_capa, bias_capa = model.layers[0].get_weights()
peso_calculado = pesos_capa[0][0]
bias_calculado = bias_capa[0]
print(f"Parámetros Finales -> Peso: {peso_calculado:.4f}, Bias: {bias_calculado:.4f}")

# --- 6. Visualización de la Curva Logística ---
plt.figure(figsize=(8, 5))

horas_planas = [x[0] for x in X_train_lista]
plt.scatter(horas_planas, y_train_lista, c=y_train_lista, cmap='coolwarm', edgecolors='k', s=60, label='Datos Reales (Excel)')

min_x = min(horas_planas) - 1
max_x = max(horas_planas) + 1
cantidad_puntos = 100
paso = (max_x - min_x) / cantidad_puntos

X_continuo_lista = [[min_x + (i * paso)] for i in range(cantidad_puntos + 1)]
# Convertimos también los datos de prueba a Tensores para poder predecir
X_continuo_tensor = tf.convert_to_tensor(X_continuo_lista)

y_probabilidad = model.predict(X_continuo_tensor, verbose=0)
X_curva_plano = [x[0] for x in X_continuo_lista]

plt.plot(X_curva_plano, y_probabilidad, color='green', linewidth=2, label='Curva de Probabilidad (Sigmoide)')

if peso_calculado != 0:
    frontera_decision = -bias_calculado / peso_calculado
    plt.axvline(x=frontera_decision, color='black', linestyle='--', alpha=0.6, label=f'Frontera 0.5 (x={frontera_decision:.2f})')

plt.title('Regresión Logística Binaria con Keras')
plt.xlabel('Horas de Estudio')
plt.ylabel('Probabilidad de Pertenencia a Clase 1')
plt.legend()
plt.grid(True, linestyle=':', alpha=0.7)
plt.show()