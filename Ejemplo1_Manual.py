"""
Documentación: Implementación de Clasificación Binaria (Perceptrón Manual) con openpyxl

Descripción:
Este script implementa un modelo de clasificación binaria iterativo, ajustando un 
'peso' y un 'bias' para trazar una frontera de decisión lineal que separe dos clases.

Gestión de Datos:
Los datos se importan desde un archivo de Excel ('Regresión.xlsx'). 
Se utiliza data_only=True para evaluar y leer el resultado de las fórmulas de Excel
en lugar de su texto original.
"""

import openpyxl
import matplotlib.pyplot as plt

# --- 1. Carga y Preparación de Datos con openpyxl ---
ruta_archivo = 'Regresión.xlsx'

# data_only=True fuerza a openpyxl a leer los valores calculados (no las fórmulas)
workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
sheet = workbook.active

horas_estudio = []
resultado_real = []

# Iterar sobre las filas del Excel, saltando encabezados (inicia en fila 2)
for row in range(2, sheet.max_row + 1):
    valor_hora = sheet.cell(row=row, column=1).value
    valor_resultado = sheet.cell(row=row, column=2).value
    
    if valor_hora is not None and valor_resultado is not None:
        try:
            hora_limpia = float(valor_hora)
            resultado_limpio = int(valor_resultado)
            
            horas_estudio.append(hora_limpia)
            resultado_real.append(resultado_limpio)
        except (ValueError, TypeError):
            continue

n_datos = len(horas_estudio)
print(f"Se han cargado {n_datos} registros exitosamente desde el Excel.")

if n_datos == 0:
    print("Error: No se pudieron cargar los datos. Verifica que el archivo no esté vacío.")
    exit()

# --- 2. Inicialización de Hiperparámetros ---
peso = 0.0      
bias = 0.0      
alpha = 0.1     
epochs = 50     

print("Iniciando entrenamiento manual...")

# --- 3. Bucle de Entrenamiento ---
for epoch in range(epochs):
    error_total = 0
    
    for i in range(n_datos):
        x = horas_estudio[i]
        y_real = resultado_real[i]
        
        # Combinación lineal
        z = peso * x + bias
        
        # Función de activación (Escalón)
        if z >= 0:
            y_pred = 1
        else:
            y_pred = 0
            
        # Cálculo del error residual
        error = y_real - y_pred
        error_total += abs(error)
        
        # Regla de actualización
        peso = peso + (alpha * error * x)
        bias = bias + (alpha * error)
    
    # Monitoreo
    if (epoch + 1) % 10 == 0:
        print(f"Epoch {epoch+1}/{epochs} | Error Acumulado: {error_total} | Peso: {peso:.4f} | Bias: {bias:.4f}")

print("\nEntrenamiento finalizado.")

# --- 4. Visualización de la Frontera de Decisión ---
plt.figure(figsize=(8, 5))
plt.scatter(horas_estudio, resultado_real, c=resultado_real, cmap='coolwarm', edgecolors='k', s=60, label='Datos Reales (Excel)')

if peso != 0:
    frontera_x = -bias / peso
    plt.axvline(x=frontera_x, color='black', linestyle='--', label=f'Frontera de Decisión (x={frontera_x:.2f})')

plt.title('Clasificación Binaria Manual (Datos desde Excel)')
plt.xlabel('Horas de Estudio')
plt.ylabel('Resultado (Clase 0 o 1)')
plt.legend()
plt.grid(True, linestyle=':', alpha=0.7)
plt.show()